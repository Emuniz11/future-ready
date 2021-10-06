# the -as- will give the package an alias making the code less tedious
import openpyxl as xl
# This import provides access to BarChart, and Reference Classes
from openpyxl.chart import BarChart, Reference

# This will load the xl workbook and return workbook object
wb = xl.load_workbook('transactions.xlsx')

# This example only has 1 sheet which can be accessed like so
sheet = wb['Sheet1']

# Now to access a particular cell on the sheet
# This will return column a row 1
# cell = sheet['a1']

# Here is another way to do this: column 1 row 1
# cell = sheet.cell(1, 1)

# The value at this cell being transaction_id (names of columns) (I'll want to skip this row)
# print(cell.value)

# Find the row length
# print(sheet.max_row)

# We can iterate through each row
# start the range at 2 instead of 1 so you don't mess with headings
for row in range(2, sheet.max_row + 1):
    # I'm manipulating the data in column 3 to change the price
    # so this is each row in column three
    cell = sheet.cell(row, 3)
    # print(cell.value) # print the rows to make sure I'm accessing the right data
    # the objective is to reduce the total price by 10% and place in a new column (column 4)
    # corrected_price is the corrected value
    corrected_price = cell.value * 0.9
    # corrected_price_cell is the cell we want the value placed into
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Using the Reference Class to select a range of values
# In this case all values from row 2 - 4 in column 4 only  ---5 total parameters---
# all of these specific values from the wb will be stored in variable values
values = Reference(sheet,
          min_row=2, max_row=sheet.max_row,
          min_col=4, max_col=4)
# decided to compare values in column 3 to column 3 since that is where the change is
values2 = Reference(sheet,
          min_row=2, max_row=sheet.max_row,
          min_col=3, max_col=3)
# Creating chart and adding values (and values2)
chart = BarChart()
chart.add_data(values)
chart.add_data(values2)
# adding chart to sheet
# e2 represent the coordinate of the top left part of the chart
sheet.add_chart(chart, 'e2')

# Now that changed have been made I need to save the workbook
# This will save to new file just in case there is a bug in the code
wb.save('transactions2.xlsx')

# This will save the changes to the original file
# wb.save('transactions.xlsx')

